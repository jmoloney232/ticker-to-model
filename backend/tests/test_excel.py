"""Phase 3 gate (owner spec): the workbook is validated by ROUND-TRIP, not by
unit inspection — generate it, recalculate it with LibreOffice headless (a
real spreadsheet engine), read the computed values back, and diff them against
the Python engine to tight tolerance. Then prove the formulas are LIVE:
change an assumption cell, recalculate, and assert the valuation moved to
exactly the engine's value under the same override.

If LibreOffice is not installed these tests SKIP LOUDLY — the phase gate has
then NOT run, and unit tests alone must not be taken as passing it."""

from __future__ import annotations

import shutil
import subprocess
from pathlib import Path

import openpyxl
import pytest
from test_cli import msft_market_provider, synthetic_provider
from test_engine import GOLDEN_VD
from test_fixtures_real import source_for

from engine.dcf import build_model
from excel.writer import write_workbook
from ingest.assemble import build_financial_history
from market.assemble import build_market_inputs

SOFFICE = (shutil.which("soffice")
           or ("/Applications/LibreOffice.app/Contents/MacOS/soffice"
               if Path("/Applications/LibreOffice.app/Contents/MacOS/soffice"
                       ).exists() else None))

pytestmark = pytest.mark.skipif(
    SOFFICE is None,
    reason="LibreOffice not found — the phase 3 round-trip GATE DID NOT RUN. "
           "Install LibreOffice; do not treat unit tests alone as the gate.")

REL_TOL = 1e-6
ABS_FLOOR = 1e-4          # for values that are legitimately ~0

# Model-sheet row key -> where the engine keeps the same number
MODEL_KEYS = {
    "revenue": ("income", "revenue"),
    "research_and_development": ("income", "research_and_development"),
    "selling_general_admin": ("income", "selling_general_admin"),
    "other_operating": ("income", "other_operating"),
    "unclassified_costs": ("income", "unclassified_costs"),
    "ebit": ("income", "operating_income"),
    "int_exp": ("income", "interest_expense"),
    "int_inc": ("income", "interest_income"),
    "pretax": ("income", "pretax_income"),
    "tax": ("income", "income_tax"),
    "ni": ("income", "net_income"),
    "cash": ("balance", "cash_and_equivalents"),
    "ar": ("balance", "accounts_receivable"),
    "inventory": ("balance", "inventory"),
    "oca": ("balance", "other_current_assets"),
    "sti": ("balance", "short_term_investments"),
    "ppe": ("balance", "ppe_net"),
    "goodwill": ("balance", "goodwill"),
    "carry": ("balance", "unattributed_carryforward"),
    "ta": ("balance", "total_assets"),
    "ap": ("balance", "accounts_payable"),
    "tl": ("balance", "total_liabilities"),
    "equity": ("balance", "stockholders_equity"),
    "pension": ("balance", "pension_liability"),
    "da": ("cashflow", "d_and_a"),
    "sbc": ("cashflow", "stock_compensation"),
    "capex": ("cashflow", "capex"),
    "wc_change": ("cashflow", "working_capital_change"),
    "cfo": ("cashflow", "cash_from_operations"),
    "dividends": ("cashflow", "dividends_paid"),
    "net_change": ("cashflow", "net_change_in_cash"),
}


def recalc(path: Path) -> openpyxl.Workbook:
    """Recalculate with LibreOffice headless and return the computed values.
    The workbook sets fullCalcOnLoad, which LibreOffice honors on conversion
    (verified by probe — a formula-only file comes back with cached values)."""
    outdir = path.parent / "recalc"
    subprocess.run(
        [SOFFICE, "--headless", "--convert-to", "xlsx",
         "--outdir", str(outdir), str(path)],
        check=True, capture_output=True, timeout=180)
    out = outdir / path.name
    assert out.exists(), "LibreOffice produced no output"
    return openpyxl.load_workbook(out, data_only=True)


def build(ticker: str, provider) -> object:
    h = build_financial_history(ticker, source_for(ticker))
    mi = build_market_inputs(ticker, provider, as_of=GOLDEN_VD)
    return build_model(h, mi, valuation_date=GOLDEN_VD)


def close(got, want) -> bool:
    if want is None:
        return got in (None, "—")
    if not isinstance(got, (int, float)):
        return False
    return abs(got - want) <= max(ABS_FLOOR, REL_TOL * abs(want))


def assert_roundtrip(m, cmap, wb) -> None:
    """Diff every substantive computed cell against the engine."""
    failures = []

    def check(key, want):
        sheet, coord = cmap[key]
        got = wb[sheet][coord].value
        if not close(got, want):
            failures.append((key, got, want))

    for i, p in enumerate(m.projections, start=1):
        for row_key, (stmt, item) in MODEL_KEYS.items():
            if f"model:{row_key}:{i}" not in cmap:
                continue                      # cost-structure-dependent rows
            src = getattr(p, stmt)
            if item not in src:
                continue
            check(f"model:{row_key}:{i}", src[item])
    for i, y in enumerate(m.ufcf, start=1):
        for key, want in (("ebit", y.ebit), ("nopat", y.nopat),
                          ("da", y.d_and_a), ("sbcadd", y.sbc_addback),
                          ("capex", y.capex), ("dnwc", y.delta_nwc),
                          ("ufcf", y.ufcf), ("t", y.exponent), ("pv", y.pv)):
            check(f"ufcf:{key}:{i}", want)

    w = m.wacc
    check("val:ke", w.cost_of_equity)
    check("val:spread", w.spread)
    check("val:kd_at", w.kd_after_tax)
    check("val:mcap", w.market_cap)
    check("val:wacc", w.wacc)
    check("val:pv_explicit", sum(y.pv for y in m.ufcf))
    if "gordon" in m.terminal:
        g = m.terminal["gordon"]
        check("val:nopat6", g.detail["nopat_n1"])
        check("val:rr", g.detail["reinvestment_rate"])
        check("val:tv_gordon", g.value_at_fyeN)
        check("val:pv_gordon", g.pv)
        check("val:ps_gordon", m.bridges["gordon"].value_per_share)
        check("val:implied_mult", m.crosschecks["implied_exit_multiple"])
    if "exit_multiple" in m.terminal:
        e = m.terminal["exit_multiple"]
        check("val:tv_exit", e.value_at_fyeN)
        check("val:pv_exit", e.pv)
        check("val:ps_exit_multiple", m.bridges["exit_multiple"].value_per_share)
        check("val:implied_g", m.crosschecks["implied_terminal_g"])
    if "epv" in m.bridges:
        fy0 = m.history.periods[-1]
        margin = m.assumptions.eff("epv_margin")
        check("val:epv_ebit", fy0.value("revenue") * margin)
        check("val:ev_epv", m.bridges["epv"].enterprise_value)
        check("val:ps_epv", m.bridges["epv"].value_per_share)
    if m.growth.available:
        check("val:growth_ps", m.growth.per_share)
        if m.growth.share_of_dcf is not None:
            check("val:growth_share", m.growth.share_of_dcf)
    bridge = next(iter(m.bridges.values()), None)
    if bridge is not None:
        by_name = {i.name: i.value for i in bridge.items}
        check("val:br_excess", by_name["excess_cash"])
        check("val:br_gd", by_name["gross_debt"])
        check("val:br_pension", by_name["pension_after_tax"])
        check("val:bridge_adj", sum(by_name.values()))

    for grid_key, sens_key in (("wacc_x_g", "g"), ("wacc_x_multiple", "mult")):
        if grid_key not in m.sensitivity:
            continue
        cells = m.sensitivity[grid_key].cells
        for i in range(len(cells)):
            for j in range(len(cells[i])):
                check(f"sens:{sens_key}:{i}:{j}", cells[i][j])

    assert wb["Model"][cmap["model:ok_bs"][1]].value == "OK"
    assert wb["Model"][cmap["model:ok_cf"][1]].value == "OK"
    assert not failures, "\n".join(
        f"{k}: excel={g!r} engine={w!r}" for k, g, w in failures[:25])


def no_error_values(wb) -> list[tuple[str, str, str]]:
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("#") \
                        and c.value.endswith(("!", "?")):
                    bad.append((ws.title, c.coordinate, c.value))
    return bad


class TestRoundTrip:
    """Owner gate #1: generated workbook, recalculated by a real engine,
    matches the Python engine cell for cell."""

    @pytest.mark.parametrize("ticker,provider_factory", [
        ("MSFT", msft_market_provider),               # clean, by_function
        ("MCD", lambda: synthetic_provider("MCD")),   # by_nature + warnings
        ("GOOGL", lambda: synthetic_provider("GOOGL")),  # share_count_derived
    ])
    def test_recalculated_workbook_matches_engine(self, tmp_path, ticker,
                                                  provider_factory):
        m = build(ticker, provider_factory())
        path = tmp_path / f"{ticker}.xlsx"
        cmap = write_workbook(m, path)
        wb = recalc(path)
        assert_roundtrip(m, cmap, wb)
        assert no_error_values(wb) == []

    def test_khc_unavailable_states_render_as_text_not_errors(self, tmp_path):
        # KHC: FY0 EBITDA <= 0 -> no exit multiple; Gordon prints its honest
        # negative-equity value (positive-but-tiny anchor, owner-accepted)
        m = build("KHC", msft_market_provider())
        path = tmp_path / "KHC.xlsx"
        cmap = write_workbook(m, path)
        wb = recalc(path)
        tv_exit = wb["Valuation"][cmap["val:tv_exit"][1]].value
        assert isinstance(tv_exit, str) and "unavailable" in tv_exit
        ps_exit = wb["Valuation"][cmap["val:ps_exit_multiple"][1]].value
        assert isinstance(ps_exit, str)
        assert close(wb["Valuation"][cmap["val:ps_gordon"][1]].value,
                     m.bridges["gordon"].value_per_share)
        assert no_error_values(wb) == []
        assert wb["Model"][cmap["model:ok_bs"][1]].value == "OK"


class TestLiveness:
    """Owner gate #2: the formulas are live — editing an assumption cell and
    recalculating moves the valuation to exactly the engine's value under the
    same override."""

    def _edited(self, tmp_path, overrides: dict):
        m = build("MSFT", msft_market_provider())
        path = tmp_path / "live.xlsx"
        cmap = write_workbook(m, path)
        wb = openpyxl.load_workbook(path)          # formulas preserved
        for name, value in overrides.items():
            sheet, coord = cmap[f"assumption:{name}"]
            wb[sheet][coord] = value
        wb.save(path)
        return m, cmap, recalc(path)

    def test_terminal_growth_edit_matches_engine_exactly(self, tmp_path):
        m, cmap, wb = self._edited(tmp_path, {"terminal_growth": 0.02})
        base = m.bridges["gordon"].value_per_share
        expected = build_model(m.history, m.market, valuation_date=GOLDEN_VD,
                               overrides={"terminal_growth": 0.02}
                               ).bridges["gordon"].value_per_share
        got = wb["Valuation"][cmap["val:ps_gordon"][1]].value
        assert got < base                           # lower g -> lower value
        assert close(got, expected)
        assert wb["Model"][cmap["model:ok_bs"][1]].value == "OK"
        assert wb["Model"][cmap["model:ok_cf"][1]].value == "OK"

    def test_cost_ratio_edit_matches_engine_exactly(self, tmp_path):
        bumped = None
        m0 = build("MSFT", msft_market_provider())
        bumped = m0.assumptions.eff("sga_pct") + 0.02
        m, cmap, wb = self._edited(tmp_path, {"sga_pct": bumped})
        expected = build_model(m.history, m.market, valuation_date=GOLDEN_VD,
                               overrides={"sga_pct": bumped}
                               ).bridges["gordon"].value_per_share
        got = wb["Valuation"][cmap["val:ps_gordon"][1]].value
        assert got < m.bridges["gordon"].value_per_share
        assert close(got, expected)

    def test_toggle_edit_is_live(self, tmp_path):
        # flipping SBC add-back ON raises UFCF and the value, exactly as the
        # engine says it should
        m, cmap, wb = self._edited(tmp_path, {"sbc_addback": True})
        expected = build_model(m.history, m.market, valuation_date=GOLDEN_VD,
                               overrides={"sbc_addback": True}
                               ).bridges["gordon"].value_per_share
        got = wb["Valuation"][cmap["val:ps_gordon"][1]].value
        assert got > m.bridges["gordon"].value_per_share
        assert close(got, expected)

    def test_seven_year_workbook_round_trips(self, tmp_path):
        # audit task 7: the sheet is laid out at the model's horizon; a 7-year
        # MSFT workbook must recalculate to the 7-year engine, cell for cell
        m0 = build("MSFT", msft_market_provider())
        m = build_model(m0.history, m0.market, valuation_date=GOLDEN_VD,
                        overrides={"forecast_years": 7})
        path = tmp_path / "msft7.xlsx"
        cmap = write_workbook(m, path)
        wb = recalc(path)
        assert_roundtrip(m, cmap, wb)

    def test_terminal_roic_fade_toggle_is_live(self, tmp_path):
        # audit task 6: the fade toggle must move the sheet exactly as the
        # engine says — ROIC_t to the midpoint with WACC, value down
        m, cmap, wb = self._edited(tmp_path, {"terminal_roic_fade": True})
        expected = build_model(m.history, m.market, valuation_date=GOLDEN_VD,
                               overrides={"terminal_roic_fade": True}
                               ).bridges["gordon"].value_per_share
        got = wb["Valuation"][cmap["val:ps_gordon"][1]].value
        assert got < m.bridges["gordon"].value_per_share
        assert close(got, expected)

    def test_midyear_toggle_asymmetry_is_live(self, tmp_path):
        # mid-year OFF: Gordon PV moves by (1+WACC)^0.5 in the exponent while
        # the exit TV keeps its full-period discount — the deliberate
        # asymmetry (worth 2–4% of value) must hold LIVE in the sheet
        m, cmap, wb = self._edited(tmp_path, {"midyear": False})
        expected = build_model(m.history, m.market, valuation_date=GOLDEN_VD,
                               overrides={"midyear": False})
        got_g = wb["Valuation"][cmap["val:ps_gordon"][1]].value
        got_x = wb["Valuation"][cmap["val:ps_exit_multiple"][1]].value
        assert close(got_g, expected.bridges["gordon"].value_per_share)
        assert close(got_x, expected.bridges["exit_multiple"].value_per_share)
        assert got_g < m.bridges["gordon"].value_per_share
        # the EPV cell shares the timing convention, so it moves too — live
        got_e = wb["Valuation"][cmap["val:ps_epv"][1]].value
        assert close(got_e, expected.bridges["epv"].value_per_share)

    def test_epv_margin_edit_is_live(self, tmp_path):
        # editing the EPV margin moves EPV and the growth line to exactly
        # the engine's values, and leaves the DCF legs untouched
        m, cmap, wb = self._edited(tmp_path, {"epv_margin": 0.30})
        expected = build_model(m.history, m.market, valuation_date=GOLDEN_VD,
                               overrides={"epv_margin": 0.30})
        got = wb["Valuation"][cmap["val:ps_epv"][1]].value
        assert close(got, expected.bridges["epv"].value_per_share)
        assert close(wb["Valuation"][cmap["val:growth_ps"][1]].value,
                     expected.growth.per_share)
        assert close(wb["Valuation"][cmap["val:ps_gordon"][1]].value,
                     m.bridges["gordon"].value_per_share)


class TestGuardsGoLiveInExcel:
    """The engine's semantic guards live in the FORMULAS: a user edit that
    breaks a precondition flips the workbook to honest text, not garbage."""

    def test_g_at_wacc_blocks_in_the_sheet(self, tmp_path):
        m = build("MSFT", msft_market_provider())
        path = tmp_path / "guard.xlsx"
        cmap = write_workbook(m, path)
        wb = openpyxl.load_workbook(path)
        sheet, coord = cmap["assumption:terminal_growth"]
        wb[sheet][coord] = 0.20                     # way above WACC (~10%)
        wb.save(path)
        wb = recalc(path)
        tv = wb["Valuation"][cmap["val:tv_gordon"][1]].value
        assert isinstance(tv, str) and "blocked" in tv
        ps = wb["Valuation"][cmap["val:ps_gordon"][1]].value
        assert isinstance(ps, str)                  # text propagates, no #VALUE!
        check = wb["Cover"][cmap["cover:check:terminal"][1]].value
        assert isinstance(check, str) and check.startswith("FAIL")
