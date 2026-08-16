"""Formula-driven workbook writer (specs/05-excel.md).

The governing rule (owner, phase 3): EVERY calculated cell contains a live
Excel formula. The only hardcoded values are input assumptions, market data,
and reported historical actuals. If a value can be computed from other cells,
it is computed in Excel, not written in by Python.

Design notes worth knowing before reading:
- Named range per assumption = the engine field name verbatim (terminal_growth,
  capex_pct, ...) plus market_price and valuation_date. One identifier across
  engine, CLI, provenance, and workbook.
- The Model sheet carries an FY0 anchor column of live =N(Historical!...)
  links (green), so the projection is visibly anchored to the reported data
  and the whole chain recalculates from the Assumptions sheet.
- Semantic guards are IN the formulas, mirroring engine behavior live: if a
  user's edits push FY5 NOPAT negative, the Gordon cells themselves flip to
  the honest "unavailable" text — same rule the engine applies at build time.
  Unavailable states are text, never Excel error values.
- The WACC × g sensitivity grid re-projects per g column (the growth path
  fades INTO g — engine semantics), via visible labeled helper blocks; the
  exit grid re-prices the base projection. No data tables (openpyxl cannot
  write them), no macros, no volatile functions, no circular references.
- Portability: plain formulas only (IF, SUMPRODUCT, POWER, MAX, ABS, N,
  ISNUMBER, TEXT) — recalculates in Excel, Google Sheets, and LibreOffice.

write_workbook() returns a cell map {logical key -> (sheet, coord)} used by
the round-trip parity test to read recalculated values back.
"""

from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from engine.assumptions import SPREAD_TABLE
from engine.dcf import G_OFFSETS, G_STEP, GRID_OFFSETS, WACC_STEP
from engine.models import ModelResult
from engine.presets import Preset

# ── formats & styles ─────────────────────────────────────────────────────────

FMT_M = "#,##0,,;(#,##0,,)"               # USD millions via comma scaling
FMT_PCT = "0.00%;(0.00%)"
FMT_X = '0.00"x"'
FMT_D = '0.0"d"'
FMT_PS = '"$"#,##0.00;("$"#,##0.00)'
FMT_DATE = "yyyy-mm-dd"
FMT_NUM = "#,##0.000"

BLUE = "0000CC"                            # hardcoded inputs / actuals
GREEN = "2E7D32"                           # pure cross-sheet links
GRAY = "808080"
RED = "9C0006"
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")

UNIT_FMT = {"rate": FMT_PCT, "ratio": FMT_PCT, "days": FMT_D, "x": FMT_X,
            "usd": FMT_PS, "shares": "#,##0,,", "flag": "General"}

# engine parity: enumerated balance-sheet membership (projections.py)
LIAB_ITEMS = ("accounts_payable", "accrued_liabilities",
              "other_current_liabilities", "deferred_revenue_current",
              "short_term_debt", "long_term_debt", "operating_lease_liability",
              "deferred_tax_liabilities", "pension_liability",
              "other_noncurrent_liabilities")
EQUITY_ITEMS = ("stockholders_equity", "noncontrolling_interest",
                "preferred_equity", "temporary_equity")
ASSET_EX_CASH_ITEMS = ("accounts_receivable", "inventory",
                       "other_current_assets", "short_term_investments",
                       "ppe_net", "goodwill", "intangibles",
                       "operating_lease_rou", "long_term_investments",
                       "investments_combined_unsplit", "other_noncurrent_assets")
FLAT_ITEMS = ("short_term_investments", "goodwill",
              "long_term_investments", "operating_lease_rou",
              "other_noncurrent_assets", "investments_combined_unsplit",
              "short_term_debt", "long_term_debt", "operating_lease_liability",
              "deferred_tax_liabilities", "pension_liability",
              "other_noncurrent_liabilities", "noncontrolling_interest",
              "preferred_equity", "temporary_equity")

HIST_IS = ("revenue", "cost_of_revenue", "gross_profit",
           "research_and_development", "selling_general_admin",
           "other_operating", "operating_income", "interest_expense",
           "interest_income", "pretax_income", "income_tax", "net_income",
           "shares_basic_wa", "shares_diluted_wa")
HIST_BS = ("cash_and_equivalents",) + ASSET_EX_CASH_ITEMS + ("total_assets",) \
          + LIAB_ITEMS + ("total_liabilities",) + EQUITY_ITEMS
HIST_CF = ("net_income", "d_and_a", "amortization_intangibles",
           "stock_compensation",
           "working_capital_change", "cash_from_operations", "capex",
           "acquisitions", "cash_from_investing", "dividends_paid",
           "buybacks", "cash_from_financing", "fx_effect",
           "net_change_in_cash")

ASSUMPTION_GROUPS = [
    ("Market data", ["market_price", "valuation_date"]),
    ("Revenue growth", ["revenue_growth_fy1", "revenue_cagr_uncapped"]),
    ("Cost structure (% of revenue, D&A-inclusive as filed)",
     ["cogs_pct", "rnd_pct", "sga_pct", "other_opex_pct",
      "unclassified_costs_pct"]),
    ("Capital intensity", ["dep_pct_beginning_ppe",
                           "amort_pct_beginning_intangibles", "da_pct_revenue",
                           "capex_pct", "capex_terminal_pct", "sbc_pct"]),
    ("Working capital", ["dso", "dio", "dpo", "oca_pct", "accrued_pct",
                         "ocl_pct", "defrev_pct"]),
    ("Taxes", ["effective_tax_fy1", "marginal_tax"]),
    ("Payout", ["payout_ratio"]),
    ("Interest", ["embedded_debt_rate", "interest_income_yield"]),
    ("Cost of capital", ["beta", "beta_raw", "erp", "risk_free",
                         "coverage_ratio"]),
    ("Horizon", ["forecast_years"]),
    ("Terminal value", ["terminal_growth", "terminal_growth_rf_ceiling",
                        "terminal_roic"]),
    ("Exit & bridge", ["exit_multiple", "share_count", "cash_floor_pct"]),
    ("Earnings power (EPV)", ["epv_margin"]),
    ("Toggles", ["midyear", "sbc_addback", "kd_synthetic", "beta_adjusted",
                 "terminal_roic_fade", "fade_curved", "capex_fade"]),
]


def _label(key: str) -> str:
    return key.replace("_", " ").capitalize()


class _Writer:
    def __init__(self, model: ModelResult, preset: Preset | None):
        self.m = model
        self.a = model.assumptions
        self.h = model.history
        self.preset = preset
        self.n_hist = len(self.h.periods)
        self.horizon = len(model.projections)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.map: dict[str, tuple[str, str]] = {}
        self.hist_row: dict[tuple[str, str], int] = {}
        self.model_row: dict[str, int] = {}
        self.val: dict[str, str] = {}      # key -> absolute ref "Valuation!$B$9"
        self.da_on_ppe = self.a.has("dep_pct_beginning_ppe")
        self.by_function = self.a.cost_structure == "by_function"

    # ── small helpers ───────────────────────────────────────────────────────
    def _name(self, name: str, sheet: str, coord: str) -> None:
        ref = f"{sheet}!${coord[0]}${coord[1:]}" if coord[1:].isdigit() \
            else f"{sheet}!{coord}"
        defn = DefinedName(name, attr_text=ref)
        try:
            self.wb.defined_names.add(defn)
        except AttributeError:             # openpyxl >= 3.1 dict interface
            self.wb.defined_names[name] = defn

    def _sheet(self, title: str):
        ws = self.wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        return ws

    def _header(self, ws, row: int, text: str, span: int = 8) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True)
        for col in range(1, span + 1):
            ws.cell(row=row, column=col).fill = HDR_FILL

    @staticmethod
    def _mcol(i: int) -> str:
        """Model/Valuation/helper column letter for projection year i (1..5);
        i=0 is the FY0 anchor column B."""
        return get_column_letter(2 + i)

    def model_ref(self, key: str, i: int) -> str:
        return f"Model!${self._mcol(i)}${self.model_row[key]}"

    def hist_ref(self, stmt: str, key: str, year_idx: int | None = None) -> str:
        row = self.hist_row[(stmt, key)]
        col = get_column_letter(2 + (self.n_hist - 1 if year_idx is None
                                     else year_idx))
        return f"Historical!${col}${row}"

    def anchor(self, stmt: str, key: str) -> str:
        """FY0 anchor formula: N() maps blank/absent history cells to 0 —
        an absent line is visible as blank on Historical, never a bare 0."""
        if (stmt, key) in self.hist_row:
            return f"N({self.hist_ref(stmt, key)})"
        return "0"

    # ── build ───────────────────────────────────────────────────────────────
    def build(self) -> None:
        # sheet creation order = tab order
        self.ws_cover = self._sheet("Cover")
        self.ws_a = self._sheet("Assumptions")
        self.ws_h = self._sheet("Historical")
        self.ws_m = self._sheet("Model")
        self.ws_v = self._sheet("Valuation")
        self.ws_s = self._sheet("Sensitivity")
        self.ws_meth = self._sheet("Methodology")

        self.write_assumptions()
        self.write_historical()
        self.write_model()
        self.write_valuation()
        self.write_sensitivity()
        self.write_methodology()
        self.write_cover()                  # last: references everything

        self.wb.calculation.fullCalcOnLoad = True

    # ── Assumptions ─────────────────────────────────────────────────────────
    def write_assumptions(self) -> None:
        ws = self.ws_a
        ws.freeze_panes = "A3"
        widths = {"A": 30, "B": 14, "C": 8, "D": 24, "E": 14, "F": 110}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws["A1"] = "Assumptions — the model's control panel"
        ws["A1"].font = Font(bold=True, size=12)
        prof = getattr(self.m.assumptions, "profile", None)
        if prof is not None:
            mm = prof.measures
            roic = (f"{mm.roic_median:+.1%}" if mm.roic_median is not None
                    else "n/a")
            cda = f"{mm.capex_da:.2f}×" if mm.capex_da is not None else "n/a"
            flag = " (REASSIGNED BY USER)" if prof.reassigned else ""
            ws["D1"] = (f"Company profile: {prof.tag}{flag} — measured: "
                        f"revenue CAGR {mm.cagr:+.1%}, latest yr "
                        f"{mm.g_latest:+.1%}, median ROIC {roic} vs WACC "
                        f"{mm.wacc:.1%}, margin range {mm.margin_range:.1%}, "
                        f"capex/D&A {cda} — rules on the Methodology sheet")
            ws["D1"].font = Font(italic=True, size=9)
        for col, head in zip("ABCDEF", ("Assumption", "Value", "Unit",
                                        "Provenance", "Derived default",
                                        "Derivation / rule"), strict=True):
            c = ws[f"{col}2"]
            c.value = head
            c.font = Font(bold=True)
            c.fill = HDR_FILL

        row = 3
        extras = {
            "market_price": (self.m.market.price.value, "usd",
                             f"market ({self.m.market.price.staleness})",
                             None, (f"Closing price as of "
                                   f"{self.m.market.price.as_of}")),
            "valuation_date": (self.m.valuation_date, "date", "input", None,
                               ("Discounting anchor — stub period measured "
                               "from FY0 fiscal year end to this date")),
        }
        for group, keys in ASSUMPTION_GROUPS:
            self._header(ws, row, group, span=6)
            row += 1
            for key in keys:
                if key in extras:
                    value, unit, prov, default, deriv = extras[key]
                elif self.a.has(key):
                    f = self.a.fields[key]
                    value, unit, prov, default = (f.effective, f.unit,
                                                  f.provenance, f.value)
                    if f.provenance == "user":
                        deriv = f"USER OVERRIDE — {f.derivation}"
                    elif f.provenance.startswith("preset"):
                        deriv = f"{f.preset_note} — {f.derivation}"
                    else:
                        deriv = f.derivation
                    if key == "beta_adjusted":
                        deriv += (" [applied at generation via the effective "
                                  "beta above — edit beta directly]")
                else:
                    continue
                ws.cell(row=row, column=1, value=_label(key))
                c = ws.cell(row=row, column=2)
                if value is not None:
                    c.value = value
                c.font = Font(color=BLUE, bold=True)
                c.fill = INPUT_FILL
                c.number_format = (FMT_DATE if unit == "date"
                                   else UNIT_FMT.get(unit, FMT_NUM))
                ws.cell(row=row, column=3, value=unit)
                ws.cell(row=row, column=4, value=prov)
                if default is not None and default != value:
                    d = ws.cell(row=row, column=5, value=default)
                    d.number_format = c.number_format
                    d.font = Font(color=GRAY)
                dv = ws.cell(row=row, column=6, value=deriv)
                dv.alignment = Alignment(wrap_text=False)
                dv.font = Font(color=GRAY, size=9)
                self._name(key, "Assumptions", f"B{row}")
                self.map[f"assumption:{key}"] = ("Assumptions", f"B{row}")
                row += 1
            row += 1

    # ── Historical ──────────────────────────────────────────────────────────
    def write_historical(self) -> None:
        ws = self.ws_h
        ws.freeze_panes = "B3"
        ws.column_dimensions["A"].width = 32
        notes_col = 2 + self.n_hist
        for i in range(self.n_hist):
            ws.column_dimensions[get_column_letter(2 + i)].width = 13
        ws.column_dimensions[get_column_letter(notes_col)].width = 50

        ws["A1"] = ("Historical financials as reported (USD millions; blue = "
                    "reported actuals, italic = ingest-derived, gray = "
                    "unmapped, treated as 0 per documented rule)")
        ws["A1"].font = Font(bold=True)
        row = 2
        ws.cell(row=row, column=1, value="Fiscal year").font = Font(bold=True)
        for i, p in enumerate(self.h.periods):
            c = ws.cell(row=row, column=2 + i, value=f"FY{p.fiscal_year}")
            c.font = Font(bold=True)
            if p.is_53_week:
                c.value += " (53w)"
        row += 1
        ws.cell(row=row, column=1, value="Fiscal year end")
        for i, p in enumerate(self.h.periods):
            c = ws.cell(row=row, column=2 + i, value=p.end)
            c.number_format = FMT_DATE
            c.font = Font(color=BLUE)
        self.hist_row[("meta", "fye")] = row
        self._name("fy0_end", "Historical",
                   f"{get_column_letter(1 + self.n_hist)}{row}")
        row += 2

        unmapped_years: dict[str, list[int]] = {}
        for w in self.h.warnings:
            if w.code == "unmapped_item" and w.item:
                unmapped_years.setdefault(w.item, []).append(w.fiscal_year)

        for stmt, title, keys in (
                ("income", "INCOME STATEMENT", HIST_IS),
                ("balance", "BALANCE SHEET", HIST_BS),
                ("cashflow", "CASH FLOW STATEMENT", HIST_CF)):
            self._header(ws, row, title, span=notes_col)
            row += 1
            for key in keys:
                facts = [getattr(p, stmt).get(key) for p in self.h.periods]
                if all(f is None for f in facts):
                    continue
                ws.cell(row=row, column=1, value=_label(key))
                sources = set()
                for i, f in enumerate(facts):
                    c = ws.cell(row=row, column=2 + i)
                    if f is None:
                        continue           # absent → blank, never a bare 0
                    c.value = f.value
                    c.number_format = FMT_M
                    if f.source == "zero_logged":
                        c.font = Font(color=GRAY, italic=True)
                    elif f.source == "derived":
                        c.font = Font(color=BLUE, italic=True)
                    else:
                        c.font = Font(color=BLUE)
                    sources.add(f.source)
                notes = []
                if key in unmapped_years:
                    ys = [y for y in unmapped_years[key] if y] or [0]
                    notes.append("unmapped in "
                                 + (f"FY{min(ys)}–FY{max(ys)}" if len(ys) > 1
                                    else f"FY{ys[0]}")
                                 + " — 0 per documented rule (see Cover)")
                elif "zero_logged" in sources:
                    notes.append("unmapped — 0 per documented rule (see Cover)")
                if "derived" in sources:
                    notes.append("derived by ingest (see Cover warnings)")
                if notes:
                    n = ws.cell(row=row, column=notes_col,
                                value="; ".join(notes))
                    n.font = Font(color=RED, size=9)
                self.hist_row[(stmt, key)] = row
                row += 1
            if stmt == "balance":
                ws.cell(row=row, column=1, value="Gross debt (ST + LT)")
                for i in range(self.n_hist):
                    col = get_column_letter(2 + i)
                    std = self.hist_row.get(("balance", "short_term_debt"))
                    ltd = self.hist_row.get(("balance", "long_term_debt"))
                    parts = [f"N({col}{r})" for r in (std, ltd) if r]
                    c = ws.cell(row=row, column=2 + i,
                                value="=" + ("+".join(parts) or "0"))
                    c.number_format = FMT_M
                self.hist_row[("balance", "gross_debt")] = row
                row += 1
            row += 1

    # ── Model ───────────────────────────────────────────────────────────────
    def _mrow(self, ws, row: int, label: str, key: str, anchor_formula,
              year_formula, fmt: str = FMT_M, bold: bool = False) -> int:
        """Register a Model row. Year formulas are lambdas evaluated in a
        SECOND pass (after every row number is known), because the statement
        chain has deliberate forward references — interest income needs the
        cash row below it, the cash plug needs the totals below it."""
        ws.cell(row=row, column=1, value=label).font = Font(bold=bold)
        self.model_row[key] = row
        self._pending.append((row, key, anchor_formula, year_formula, fmt,
                              bold))
        for i in range(1, self.horizon + 1):
            self.map[f"model:{key}:{i}"] = ("Model", f"{self._mcol(i)}{row}")
        return row + 1

    def _flush_model_rows(self, ws) -> None:
        for row, _key, anchor_formula, year_formula, fmt, bold in self._pending:
            if callable(anchor_formula):
                anchor_formula = anchor_formula()
            if anchor_formula is not None:
                c = ws.cell(row=row, column=2, value=anchor_formula)
                c.number_format = fmt
                c.font = Font(color=GREEN, bold=bold)
            for i in range(1, self.horizon + 1):
                c = ws.cell(row=row, column=2 + i, value=year_formula(i))
                c.number_format = fmt
                c.font = Font(bold=bold)
        self._pending = []

    def write_model(self) -> None:
        ws = self.ws_m
        self._pending: list = []
        ws.freeze_panes = "C3"
        ws.column_dimensions["A"].width = 32
        for i in range(self.horizon + 1):
            ws.column_dimensions[self._mcol(i)].width = 13
        fy0 = self.h.periods[-1].fiscal_year
        ws["A1"] = ("Projected three-statement model (USD millions) — every "
                    "cell is a live formula; FY0 column links Historical "
                    "(green); drivers come from the Assumptions named ranges")
        ws["A1"].font = Font(bold=True)
        ws.cell(row=2, column=1, value="Fiscal year").font = Font(bold=True)
        ws.cell(row=2, column=2, value=f"FY{fy0} (anchor)").font = Font(bold=True)
        for i in range(1, self.horizon + 1):
            ws.cell(row=2, column=2 + i,
                    value=f"FY{fy0 + i}E").font = Font(bold=True)

        r = 4
        self._header(ws, r, "DRIVERS", span=2 + self.horizon)
        r += 1
        # Fade shape is LIVE: linear, or half-cosine when fade_curved is on
        # (compounder profile) — same closed form as engine growth_path
        n1 = self.horizon - 1
        r = self._mrow(ws, r, "Revenue growth (fade to terminal g)", "growth",
                       None, lambda i: (
                           f"=IF(fade_curved,terminal_growth+"
                           f"(revenue_growth_fy1-terminal_growth)*"
                           f"(1+COS(PI()*{i - 1}/{n1}))/2,"
                           f"revenue_growth_fy1+{i - 1}/{n1}*"
                           f"(terminal_growth-revenue_growth_fy1))"), FMT_PCT)
        r = self._mrow(ws, r, "Tax rate (effective → marginal fade)", "tax_rate",
                       None, lambda i: f"=effective_tax_fy1+{i - 1}/"
                                       f"{self.horizon - 1}*(marginal_tax"
                                       f"-effective_tax_fy1)", FMT_PCT)
        gd = f"N({self.hist_ref('balance', 'gross_debt')})"
        r = self._mrow(ws, r, "Gross debt (held constant — no revolver in v1)",
                       "debt", f"={gd}",
                       lambda i: f"=$B${self.model_row['debt']}")
        r += 1

        self._header(ws, r, "INCOME STATEMENT", span=2 + self.horizon)
        r += 1
        R = self.model_row
        r = self._mrow(ws, r, "Revenue", "revenue",
                       f"={self.anchor('income', 'revenue')}",
                       lambda i: f"={self._mcol(i - 1)}{R['revenue']}"
                                 f"*(1+{self._mcol(i)}{R['growth']})",
                       bold=True)
        cost_keys = []
        if self.by_function:
            r = self._mrow(ws, r, "Cost of revenue", "cost_of_revenue",
                           f"={self.anchor('income', 'cost_of_revenue')}",
                           lambda i: f"=cogs_pct*{self._mcol(i)}{R['revenue']}")
            cost_keys.append("cost_of_revenue")
            r = self._mrow(ws, r, "Gross profit", "gross_profit",
                           None,
                           lambda i: f"={self._mcol(i)}{R['revenue']}"
                                     f"-{self._mcol(i)}{R['cost_of_revenue']}")
        for label, key, pct in (
                ("Research & development", "research_and_development", "rnd_pct"),
                ("Selling, general & admin", "selling_general_admin", "sga_pct"),
                ("Other operating", "other_operating", "other_opex_pct"),
                ("Unclassified costs (margin-identity closure)",
                 "unclassified_costs", "unclassified_costs_pct")):
            r = self._mrow(ws, r, label, key,
                           f"={self.anchor('income', key)}"
                           if key != "unclassified_costs" else None,
                           lambda i, p=pct: f"={p}*{self._mcol(i)}{R['revenue']}")
            cost_keys.append(key)
        r = self._mrow(ws, r, "EBIT", "ebit",
                       f"={self.anchor('income', 'operating_income')}",
                       lambda i: "=" + self._mcol(i) + str(R["revenue"]) + "-"
                       + "-".join(f"{self._mcol(i)}{R[k]}" for k in cost_keys),
                       bold=True)
        r = self._mrow(ws, r, "Interest expense (beginning debt)", "int_exp",
                       None, lambda i: f"=embedded_debt_rate"
                                       f"*$B${R['debt']}")
        r = self._mrow(ws, r, "Interest income (beginning cash + STI)",
                       "int_inc", None,
                       lambda i: f"=interest_income_yield*({self._mcol(i - 1)}"
                                 f"{R['cash']}+{self._mcol(i - 1)}{R['sti']})")
        r = self._mrow(ws, r, "Pretax income", "pretax", None,
                       lambda i: f"={self._mcol(i)}{R['ebit']}"
                                 f"-{self._mcol(i)}{R['int_exp']}"
                                 f"+{self._mcol(i)}{R['int_inc']}")
        r = self._mrow(ws, r, "Income tax", "tax", None,
                       lambda i: f"={self._mcol(i)}{R['tax_rate']}"
                                 f"*{self._mcol(i)}{R['pretax']}")
        r = self._mrow(ws, r, "Net income", "ni", None,
                       lambda i: f"={self._mcol(i)}{R['pretax']}"
                                 f"-{self._mcol(i)}{R['tax']}", bold=True)
        r += 1

        self._header(ws, r, "MEMO — D&A, SBC, CAPEX (spec 04: D&A placement)",
                     span=2 + self.horizon)
        r += 1
        # Split D&A basis (owner-approved 2026-08-16): depreciation drives the
        # PP&E roll (MIN = identity floor — net PP&E cannot fall below zero);
        # intangible amortization runs the intangibles balance off and stops
        # at zero. D&A memo = the sum, exactly as the engine computes it.
        if self.da_on_ppe:
            dep = lambda i: (f"=MIN(dep_pct_beginning_ppe"
                             f"*{self._mcol(i - 1)}{R['ppe']},"
                             f"{self._mcol(i - 1)}{R['ppe']}"
                             f"+{self._mcol(i)}{R['capex']})")
        else:
            dep = lambda i: f"=da_pct_revenue*{self._mcol(i)}{R['revenue']}"
        r = self._mrow(ws, r, "Depreciation (memo — PP&E roll)", "dep",
                       None, dep)
        if self.da_on_ppe and self.a.has("amort_pct_beginning_intangibles"):
            amort = lambda i: (f"=MIN(amort_pct_beginning_intangibles"
                               f"*{self._mcol(i - 1)}{R['intangibles']},"
                               f"{self._mcol(i - 1)}{R['intangibles']})")
            amort_label = "Intangible amortization (memo — run-off)"
        else:
            amort = lambda i: 0
            amort_label = ("Intangible amortization (unobservable — combined "
                           "basis above)" if self.da_on_ppe
                           else "Intangible amortization (n/a)")
        r = self._mrow(ws, r, amort_label, "amort", None, amort)
        r = self._mrow(ws, r, "D&A (memo — CF add-back, EBITDA only)",
                       "da", None,
                       lambda i: f"={self._mcol(i)}{R['dep']}"
                                 f"+{self._mcol(i)}{R['amort']}")
        r = self._mrow(ws, r, "Stock-based compensation", "sbc", None,
                       lambda i: f"=sbc_pct*{self._mcol(i)}{R['revenue']}")
        # Capex % of revenue is LIVE: flat, or fading to maintenance when
        # capex_fade is on (reinvestment-heavy profile) — mirrors capex_path
        r = self._mrow(ws, r, "Capex", "capex", None,
                       lambda i: (
                           f"=IF(capex_fade,capex_pct+{i - 1}/"
                           f"{self.horizon - 1}*(capex_terminal_pct-capex_pct),"
                           f"capex_pct)*{self._mcol(i)}{R['revenue']}"))
        r += 1

        # NOTE: cash row must exist before int_inc formulas resolve — rows are
        # written in statement order; Excel forward references are fine.
        self._header(ws, r, "BALANCE SHEET — ASSETS", span=2 + self.horizon)
        r += 1
        r = self._mrow(ws, r, "Cash & equivalents (THE PLUG)", "cash",
                       f"={self.anchor('balance', 'cash_and_equivalents')}",
                       lambda i: f"={self._mcol(i)}{R['tl']}"
                                 f"+{self._mcol(i)}{R['eq_side']}"
                                 f"-{self._mcol(i)}{R['assets_ex_cash']}",
                       bold=True)
        basis = (lambda i: f"{self._mcol(i)}{R['cost_of_revenue']}") \
            if self.by_function else \
            (lambda i: f"({self._mcol(i)}{R['revenue']}"
                       f"-{self._mcol(i)}{R['ebit']})")
        r = self._mrow(ws, r, "Accounts receivable (DSO)", "ar",
                       f"={self.anchor('balance', 'accounts_receivable')}",
                       lambda i: f"=dso/365*{self._mcol(i)}{R['revenue']}")
        r = self._mrow(ws, r, "Inventory (DIO)", "inventory",
                       f"={self.anchor('balance', 'inventory')}",
                       lambda i: f"=dio/365*{basis(i)}")
        r = self._mrow(ws, r, "Other current assets", "oca",
                       f"={self.anchor('balance', 'other_current_assets')}",
                       lambda i: f"=oca_pct*{self._mcol(i)}{R['revenue']}")
        r = self._mrow(ws, r, "Short-term investments (held flat)", "sti",
                       f"={self.anchor('balance', 'short_term_investments')}",
                       lambda i: f"=$B${R['sti']}")
        r = self._mrow(ws, r, "Net PP&E (roll-forward)", "ppe",
                       f"={self.anchor('balance', 'ppe_net')}",
                       lambda i: f"={self._mcol(i - 1)}{R['ppe']}"
                                 f"+{self._mcol(i)}{R['capex']}"
                                 f"-{self._mcol(i)}{R['dep']}")
        r = self._mrow(ws, r, "Intangibles (run-off by amortization)",
                       "intangibles",
                       f"={self.anchor('balance', 'intangibles')}",
                       lambda i: f"={self._mcol(i - 1)}{R['intangibles']}"
                                 f"-{self._mcol(i)}{R['amort']}")
        flat_assets = [("Goodwill (held flat)", "goodwill", "goodwill"),
                       ("Operating lease ROU (held flat)", "rou",
                        "operating_lease_rou"),
                       ("Long-term investments (held flat)", "lti",
                        "long_term_investments"),
                       ("Combined unsplit investments (held flat)", "unsplit",
                        "investments_combined_unsplit"),
                       ("Other noncurrent assets (held flat)", "onca",
                        "other_noncurrent_assets")]
        for label, key, item in flat_assets:
            r = self._mrow(ws, r, label, key,
                           f"={self.anchor('balance', item)}",
                           lambda i, k=key: f"=$B${R[k]}")
        aex = ("ar", "inventory", "oca", "sti", "ppe", "goodwill",
               "intangibles", "rou", "lti", "unsplit", "onca")
        r = self._mrow(ws, r, "Assets excl. cash", "assets_ex_cash", None,
                       lambda i: "=" + "+".join(f"{self._mcol(i)}{R[k]}"
                                                for k in aex))
        r = self._mrow(ws, r, "TOTAL ASSETS", "ta", None,
                       lambda i: f"={self._mcol(i)}{R['assets_ex_cash']}"
                                 f"+{self._mcol(i)}{R['cash']}", bold=True)
        r += 1

        self._header(ws, r, "BALANCE SHEET — LIABILITIES & EQUITY",
                     span=2 + self.horizon)
        r += 1
        r = self._mrow(ws, r, "Accounts payable (DPO)", "ap",
                       f"={self.anchor('balance', 'accounts_payable')}",
                       lambda i: f"=dpo/365*{basis(i)}")
        for label, key, pct in (
                ("Accrued liabilities", "accrued", "accrued_pct"),
                ("Other current liabilities", "ocl", "ocl_pct"),
                ("Deferred revenue (current)", "defrev", "defrev_pct")):
            r = self._mrow(ws, r, label, key,
                           f"={self.anchor('balance', 'accrued_liabilities' if key == 'accrued' else 'other_current_liabilities' if key == 'ocl' else 'deferred_revenue_current')}",
                           lambda i, p=pct: f"={p}*{self._mcol(i)}{R['revenue']}")
        flat_liabs = [("Short-term debt (held flat)", "std", "short_term_debt"),
                      ("Long-term debt (held flat)", "ltd", "long_term_debt"),
                      ("Operating lease liability (held flat)", "lease",
                       "operating_lease_liability"),
                      ("Deferred tax liabilities (held flat)", "dtl",
                       "deferred_tax_liabilities"),
                      ("Pension liability (held flat)", "pension",
                       "pension_liability"),
                      ("Other noncurrent liabilities (held flat)", "oncl",
                       "other_noncurrent_liabilities")]
        for label, key, item in flat_liabs:
            r = self._mrow(ws, r, label, key,
                           f"={self.anchor('balance', item)}",
                           lambda i, k=key: f"=$B${R[k]}")
        liab_keys = ("ap", "accrued", "ocl", "defrev", "std", "ltd", "lease",
                     "dtl", "pension", "oncl")
        eq_anchor_keys = [("balance", k) for k in EQUITY_ITEMS]
        carry_anchor = ("=" + f"$B${R['cash']}" + "-("
                        + "+".join(f"$B${R[k]}" for k in liab_keys)
                        + "+" + "+".join(f"N({self.hist_ref(s, k)})"
                                         for s, k in eq_anchor_keys
                                         if (s, k) in self.hist_row)
                        + "-(" + "+".join(f"$B${R[k]}" for k in aex) + "))")
        r = self._mrow(ws, r, "Unattributed carryforward (FY0 mapping residual,"
                              " held flat — spec 04)", "carry",
                       carry_anchor, lambda i: f"=$B${R['carry']}")
        r = self._mrow(ws, r, "TOTAL LIABILITIES", "tl", None,
                       lambda i: "=" + "+".join(f"{self._mcol(i)}{R[k]}"
                                                for k in liab_keys)
                                 + f"+{self._mcol(i)}{R['carry']}", bold=True)
        r = self._mrow(ws, r, "Stockholders' equity (roll-forward)", "equity",
                       f"={self.anchor('balance', 'stockholders_equity')}",
                       lambda i: f"={self._mcol(i - 1)}{R['equity']}"
                                 f"+{self._mcol(i)}{R['ni']}"
                                 f"-{self._mcol(i)}{R['dividends']}"
                                 f"+{self._mcol(i)}{R['sbc']}")
        for label, key, item in (
                ("Noncontrolling interest (held flat)", "nci",
                 "noncontrolling_interest"),
                ("Preferred equity (held flat)", "pref", "preferred_equity"),
                ("Temporary equity (held flat)", "temp", "temporary_equity")):
            r = self._mrow(ws, r, label, key,
                           f"={self.anchor('balance', item)}",
                           lambda i, k=key: f"=$B${R[k]}")
        r = self._mrow(ws, r, "Equity side (incl. NCI, preferred, temporary)",
                       "eq_side", None,
                       lambda i: f"={self._mcol(i)}{R['equity']}"
                                 f"+{self._mcol(i)}{R['nci']}"
                                 f"+{self._mcol(i)}{R['pref']}"
                                 f"+{self._mcol(i)}{R['temp']}")
        r = self._mrow(ws, r, "TOTAL LIABILITIES & EQUITY", "tle", None,
                       lambda i: f"={self._mcol(i)}{R['tl']}"
                                 f"+{self._mcol(i)}{R['eq_side']}", bold=True)
        r += 1

        self._header(ws, r, "CASH FLOW STATEMENT (indirect)",
                     span=2 + self.horizon)
        r += 1
        r = self._mrow(ws, r, "Net income", "cf_ni", None,
                       lambda i: f"={self._mcol(i)}{R['ni']}")
        r = self._mrow(ws, r, "D&A", "cf_da", None,
                       lambda i: f"={self._mcol(i)}{R['da']}")
        r = self._mrow(ws, r, "Stock-based compensation", "cf_sbc", None,
                       lambda i: f"={self._mcol(i)}{R['sbc']}")
        nwc_terms = ("ar", "inventory", "oca")
        nwc_neg = ("ap", "accrued", "ocl", "defrev")
        r = self._mrow(ws, r, "Operating NWC (memo)", "nwc",
                       "=" + "+".join(f"$B${R[k]}" for k in nwc_terms)
                       + "-" + "-".join(f"$B${R[k]}" for k in nwc_neg),
                       lambda i: "=" + "+".join(f"{self._mcol(i)}{R[k]}"
                                                for k in nwc_terms)
                       + "-" + "-".join(f"{self._mcol(i)}{R[k]}"
                                        for k in nwc_neg))
        r = self._mrow(ws, r, "Working capital change (−ΔNWC)", "wc_change",
                       None,
                       lambda i: f"=-({self._mcol(i)}{R['nwc']}"
                                 f"-{self._mcol(i - 1)}{R['nwc']})")
        r = self._mrow(ws, r, "Cash from operations", "cfo", None,
                       lambda i: f"={self._mcol(i)}{R['cf_ni']}"
                                 f"+{self._mcol(i)}{R['cf_da']}"
                                 f"+{self._mcol(i)}{R['cf_sbc']}"
                                 f"+{self._mcol(i)}{R['wc_change']}",
                       bold=True)
        r = self._mrow(ws, r, "Capex", "cf_capex", None,
                       lambda i: f"=-{self._mcol(i)}{R['capex']}")
        r = self._mrow(ws, r, "Cash from investing", "cfi", None,
                       lambda i: f"={self._mcol(i)}{R['cf_capex']}", bold=True)
        r = self._mrow(ws, r, "Dividends paid", "dividends", None,
                       lambda i: f"=payout_ratio"
                                 f"*MAX(0,{self._mcol(i)}{R['ni']})")
        r = self._mrow(ws, r, "Cash from financing", "cff", None,
                       lambda i: f"=-{self._mcol(i)}{R['dividends']}",
                       bold=True)
        r = self._mrow(ws, r, "Net change in cash", "net_change", None,
                       lambda i: f"={self._mcol(i)}{R['cfo']}"
                                 f"+{self._mcol(i)}{R['cfi']}"
                                 f"+{self._mcol(i)}{R['cff']}", bold=True)
        r += 1

        self._header(ws, r, "LIVE CHECKS — recalculate when anything changes",
                     span=2 + self.horizon)
        r += 1
        r = self._mrow(ws, r, "Balance sheet tie (assets − L&E; must be 0)",
                       "check_bs", None,
                       lambda i: f"={self._mcol(i)}{R['ta']}"
                                 f"-{self._mcol(i)}{R['tle']}", FMT_NUM)
        r = self._mrow(ws, r, "Cash flow tie (net change − Δcash; must be 0)",
                       "check_cf", None,
                       lambda i: f"={self._mcol(i)}{R['net_change']}"
                                 f"-({self._mcol(i)}{R['cash']}"
                                 f"-{self._mcol(i - 1)}{R['cash']})", FMT_NUM)
        rng = lambda key: (f"C{R[key]}:"
                           f"{self._mcol(self.horizon)}{R[key]}")
        for label, key, formula in (
                ("BALANCE SHEET TIES", "ok_bs",
                 (f'=IF(SUMPRODUCT(ABS({rng("check_bs")}))<1,"OK",'
                 f'"FAIL — see row {R["check_bs"]}")')),
                ("CASH FLOW TIES", "ok_cf",
                 (f'=IF(SUMPRODUCT(ABS({rng("check_cf")}))<1,"OK",'
                 f'"FAIL — see row {R["check_cf"]}")'))):
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            c = ws.cell(row=r, column=3, value=formula)
            c.font = Font(bold=True)
            self.model_row[key] = r
            self.map[f"model:{key}"] = ("Model", f"C{r}")
            r += 1
        self._flush_model_rows(ws)

    # ── Valuation ───────────────────────────────────────────────────────────
    def _vrow(self, ws, row: int, label: str, key: str, formula,
              fmt: str = FMT_M, bold: bool = False, col: int = 2):
        ws.cell(row=row, column=1, value=label).font = Font(bold=bold)
        c = ws.cell(row=row, column=col, value=formula)
        c.number_format = fmt
        c.font = Font(bold=bold)
        letter = get_column_letter(col)
        self.val[key] = f"Valuation!${letter}${row}"
        self.map[f"val:{key}"] = ("Valuation", f"{letter}{row}")
        return row + 1

    # ── per-method formula blocks (see BLOCK_BUILDERS in write_valuation) ───
    def _block_gordon(self, ws, r: int) -> tuple[int, str]:
        V = self.val
        self._header(ws, r, "TERMINAL VALUE — GORDON (reinvestment-consistent)",
                     span=8)
        r += 1
        ebit5 = f"{self._mcol(self.horizon)}{self.val['u_ebit']}"
        r = self._vrow(ws, r, "NOPAT (N+1) = EBIT5 × (1+g) × (1−marginal)",
                       "nopat6",
                       f"={ebit5}*(1+terminal_growth)*(1-marginal_tax)")
        r = self._vrow(ws, r, "ROIC used (derived; falls back to WACC when "
                              "unavailable or ≤ g; terminal_roic_fade → "
                              "midpoint with WACC)", "roic",
                       f'=IF(OR(terminal_roic="",terminal_roic<='
                       f"terminal_growth),{V['wacc']},"
                       f"IF(terminal_roic_fade,(terminal_roic+{V['wacc']})/2,"
                       "terminal_roic))", FMT_PCT)
        r = self._vrow(ws, r, "Reinvestment rate RR = g / ROIC", "rr",
                       f"=terminal_growth/{V['roic']}", FMT_PCT)
        guard = (f'IF({V["nopat6"]}<=0,"unavailable — negative terminal NOPAT '
                 f'anchor (see Methodology)",IF(terminal_growth>={V["wacc"]},'
                 f'"blocked — terminal g must be below WACC",{{}}))')
        r = self._vrow(ws, r, "TV at FYE5", "tv_gordon",
                       "=" + guard.format(
                           f"{V['nopat6']}*(1-{V['rr']})"
                           f"/({V['wacc']}-terminal_growth)"))
        r = self._vrow(ws, r, "PV of Gordon TV", "pv_gordon",
                       f"=IF(ISNUMBER({V['tv_gordon']}),{V['tv_gordon']}"
                       f"*POWER(1+{V['wacc']},-{V['exp_gordon']}),"
                       f"{V['tv_gordon']})")
        r += 1
        ev = (f"=IF(ISNUMBER({V['pv_gordon']}),{V['pv_explicit']}"
              f"+{V['pv_gordon']},{V['pv_gordon']})")
        return r, ev

    def _block_exit(self, ws, r: int) -> tuple[int, str]:
        V = self.val
        self._header(ws, r, "TERMINAL VALUE — EXIT MULTIPLE", span=8)
        r += 1
        ebit5 = f"{self._mcol(self.horizon)}{self.val['u_ebit']}"
        da5 = f"{self._mcol(self.horizon)}{self.val['u_da']}"
        r = self._vrow(ws, r, "EBITDA (FY5) = EBIT5 + memo D&A5", "ebitda5",
                       f"={ebit5}+{da5}")
        r = self._vrow(ws, r, "TV at FYE5 (multiple × EBITDA5)", "tv_exit",
                       f'=IF(exit_multiple="","unavailable — no exit multiple '
                       f'(FY0 EBITDA ≤ 0)",IF({V["ebitda5"]}<=0,'
                       f'"unavailable — negative FY5 EBITDA",'
                       f"exit_multiple*{V['ebitda5']}))")
        r = self._vrow(ws, r, "PV of exit TV (full tN — deliberate asymmetry)",
                       "pv_exit",
                       f"=IF(ISNUMBER({V['tv_exit']}),{V['tv_exit']}"
                       f"*POWER(1+{V['wacc']},-{V['tn']}),{V['tv_exit']})")
        r += 1
        ev = (f"=IF(ISNUMBER({V['pv_exit']}),{V['pv_explicit']}"
              f"+{V['pv_exit']},{V['pv_exit']})")
        return r, ev

    def _block_epv(self, ws, r: int) -> tuple[int, str]:
        """Earnings power: normalized EBIT (FY0 revenue × epv_margin, the
        profile-ruled editable assumption), marginal-taxed, capitalized at
        WACC with the SAME first-flow timing as the explicit years — a flat
        perpetuity the g = 0 DCF converges to exactly (tested)."""
        V = self.val
        self._header(ws, r, "EARNINGS POWER — NO GROWTH (EPV)", span=8)
        r += 1
        rev0 = f"N({self.hist_ref('income', 'revenue')})"
        r = self._vrow(ws, r, "Normalized EBIT = FY0 revenue × EPV margin",
                       "epv_ebit", f"={rev0}*epv_margin")
        r = self._vrow(ws, r, "Normalized NOPAT (marginal-taxed — same rate "
                              "as terminal NOPAT)", "epv_nopat",
                       f"={V['epv_ebit']}*(1-marginal_tax)")
        t1 = f"C{self.val['u_t']}"
        r = self._vrow(ws, r, "EPV enterprise value = NOPAT / WACC × "
                              "(1+WACC)^(1−t₁) (maintenance capex = D&A, "
                              "flat working capital — D&A cancels)",
                       "ev_epv",
                       f'=IF({V["epv_ebit"]}<=0,"unavailable — negative '
                       f'normalized earnings (see Methodology)",'
                       f"{V['epv_nopat']}/{V['wacc']}"
                       f"*POWER(1+{V['wacc']},1-{t1}))")
        r += 1
        return r, f"={V['ev_epv']}"

    def write_valuation(self) -> None:
        ws = self.ws_v
        ws.freeze_panes = "B3"
        ws.column_dimensions["A"].width = 40
        for i in range(1, 9):
            ws.column_dimensions[get_column_letter(1 + i)].width = 13
        ws["A1"] = ("Valuation — WACC build-up, UFCF, one block per valuation "
                    "method, EV → equity bridge (all live formulas)")
        ws["A1"].font = Font(bold=True)
        r = 3
        V = self.val

        self._header(ws, r, "WACC BUILD-UP", span=8)
        r += 1
        r = self._vrow(ws, r, "Cost of equity = rf + β × ERP", "ke",
                       "=risk_free+beta*erp", FMT_PCT)
        # innermost else = the table's own bottom bracket (never hardcoded —
        # a truncated distressed range was the audit's headline bug)
        spread_if = f"{SPREAD_TABLE[-1][2]}"
        rating_if = f'"{SPREAD_TABLE[-1][1]}"'
        for floor, rating, spread in reversed(SPREAD_TABLE[:-1]):
            spread_if = f"IF(coverage_ratio>{floor},{spread},{spread_if})"
            rating_if = f'IF(coverage_ratio>{floor},"{rating}",{rating_if})'
        top = SPREAD_TABLE[0]
        r = self._vrow(ws, r, "Synthetic rating (coverage → bracket)", "rating",
                       f'=IF(coverage_ratio="","{top[1]} (no traceable '
                       f'interest)",{rating_if})', "General")
        r = self._vrow(ws, r, "Default spread", "spread",
                       f'=IF(coverage_ratio="",{top[2]},{spread_if})', FMT_PCT)
        r = self._vrow(ws, r, "Kd pre-tax (toggle: synthetic | embedded)",
                       "kd_pre",
                       f"=IF(kd_synthetic,risk_free+{V['spread']},"
                       f"IF(embedded_debt_rate>0,embedded_debt_rate,"
                       f"risk_free+{V['spread']}))", FMT_PCT)
        r = self._vrow(ws, r, "Kd after tax (marginal — same rate as terminal "
                              "NOPAT)", "kd_at",
                       f"={V['kd_pre']}*(1-marginal_tax)", FMT_PCT)
        r = self._vrow(ws, r, "Market cap (price × shares)", "mcap",
                       "=market_price*share_count")
        r = self._vrow(ws, r, "Gross debt (book — net debt appears only in "
                              "the bridge)", "debt",
                       f"=N({self.hist_ref('balance', 'gross_debt')})")
        r = self._vrow(ws, r, "Weight of equity", "we",
                       f"={V['mcap']}/({V['mcap']}+{V['debt']})", FMT_PCT)
        r = self._vrow(ws, r, "WACC", "wacc",
                       f"={V['we']}*{V['ke']}+(1-{V['we']})*{V['kd_at']}",
                       FMT_PCT, bold=True)
        r += 1

        self._header(ws, r, "DISCOUNTING", span=8)
        r += 1
        r = self._vrow(ws, r, "Stub (years elapsed since FY0 end)", "stub",
                       "=(valuation_date-fy0_end)/365.25", FMT_NUM)
        r = self._vrow(ws, r, "Exit exponent tN (full period — a sale is a "
                              "year-end event)", "tn",
                       f"={self.horizon}-{V['stub']}", FMT_NUM)
        r = self._vrow(ws, r, "Gordon exponent (mid-year toggle)", "exp_gordon",
                       f"={V['tn']}-IF(midyear,0.5,0)", FMT_NUM)
        r += 1

        self._header(ws, r, "UNLEVERED FREE CASH FLOW", span=8)
        r += 1
        fy0 = self.h.periods[-1].fiscal_year
        ws.cell(row=r, column=1, value="Fiscal year").font = Font(bold=True)
        for i in range(1, self.horizon + 1):
            ws.cell(row=r, column=2 + i,
                    value=f"FY{fy0 + i}E").font = Font(bold=True)
        r += 1

        def urow(row, label, key, formula, fmt=FMT_M, bold=False):
            ws.cell(row=row, column=1, value=label).font = Font(bold=bold)
            for i in range(1, self.horizon + 1):
                c = ws.cell(row=row, column=2 + i, value=formula(i))
                c.number_format = fmt
                c.font = Font(color=GREEN) if formula(i).startswith("=Model!") \
                    and "+" not in formula(i) and "*" not in formula(i) \
                    else Font(bold=bold)
                self.map[f"ufcf:{key}:{i}"] = ("Valuation",
                                               f"{self._mcol(i)}{row}")
            self.val[f"u_{key}"] = row
            return row + 1

        r = urow(r, "EBIT", "ebit", lambda i: f"={self.model_ref('ebit', i)}")
        r = urow(r, "Tax rate", "taxrate",
                 lambda i: f"={self.model_ref('tax_rate', i)}", FMT_PCT)
        r = urow(r, "NOPAT", "nopat",
                 lambda i: f"={self._mcol(i)}{self.val['u_ebit']}"
                           f"*(1-{self._mcol(i)}{self.val['u_taxrate']})")
        r = urow(r, "D&A", "da", lambda i: f"={self.model_ref('da', i)}")
        r = urow(r, "SBC add-back (toggle; default expensed)", "sbcadd",
                 lambda i: f"=IF(sbc_addback,{self.model_ref('sbc', i)},0)")
        r = urow(r, "Capex", "capex",
                 lambda i: f"={self.model_ref('capex', i)}")
        r = urow(r, "ΔNWC", "dnwc",
                 lambda i: f"=-{self.model_ref('wc_change', i)}")
        r = urow(r, "UFCF", "ufcf",
                 lambda i: f"={self._mcol(i)}{self.val['u_nopat']}"
                           f"+{self._mcol(i)}{self.val['u_da']}"
                           f"+{self._mcol(i)}{self.val['u_sbcadd']}"
                           f"-{self._mcol(i)}{self.val['u_capex']}"
                           f"-{self._mcol(i)}{self.val['u_dnwc']}", FMT_M, True)
        r = urow(r, "Discount exponent t", "t",
                 lambda i: f"={i}-{V['stub']}-IF(midyear,0.5,0)", FMT_NUM)
        r = urow(r, "PV of UFCF", "pv",
                 lambda i: f"={self._mcol(i)}{self.val['u_ufcf']}"
                           f"*POWER(1+{V['wacc']},"
                           f"-{self._mcol(i)}{self.val['u_t']})")
        self.ufcf_rng = (f"Valuation!$C${self.val['u_ufcf']}:"
                         f"${self._mcol(self.horizon)}${self.val['u_ufcf']}")
        self.t_rng = (f"Valuation!$C${self.val['u_t']}:"
                      f"${self._mcol(self.horizon)}${self.val['u_t']}")
        r = self._vrow(ws, r, "PV of explicit years (sum)", "pv_explicit",
                       f"=SUM(C{self.val['u_pv']}:"
                       f"{self._mcol(self.horizon)}{self.val['u_pv']})",
                       FMT_M, bold=True)
        r += 1

        # ── method blocks — BLOCK_BUILDERS keyed by registry id ─────────────
        # The frame below (bridge columns, equity, per-share) iterates the
        # registry generically; the formula core of each method necessarily
        # lives in its own builder (live formulas ARE the method). Adding a
        # method = one engine entry + one builder here; an id with no builder
        # fails loudly rather than silently dropping a column.
        builders = {"gordon": self._block_gordon,
                    "exit_multiple": self._block_exit,
                    "epv": self._block_epv}
        methods = sorted(self.m.methods, key=lambda mr: mr.order)
        ev_formula: dict[str, str] = {}
        for mr in methods:
            if mr.id not in builders:
                raise KeyError(
                    f"no workbook block builder for valuation method "
                    f"'{mr.id}' — every registry entry must contribute "
                    "its formula block")
            r, ev_formula[mr.id] = builders[mr.id](ws, r)

        self._header(ws, r, "IMPLIED CROSS-CHECKS", span=8)
        r += 1
        r = self._vrow(ws, r, "Your terminal g implies an exit multiple of",
                       "implied_mult",
                       f"=IF(AND(ISNUMBER({V['tv_gordon']}),{V['ebitda5']}>0),"
                       f"{V['tv_gordon']}/{V['ebitda5']},\"n/a\")", FMT_X)
        r = self._vrow(ws, r, "Your exit multiple implies terminal g of",
                       "implied_g",
                       f"=IF(ISNUMBER({V['tv_exit']}),{V['wacc']}"
                       f"-{V['nopat6']}*(1-{V['rr']})/{V['tv_exit']},\"n/a\")",
                       FMT_PCT)
        r += 1

        self._header(ws, r, "EV → EQUITY BRIDGE", span=8)
        r += 1
        # one column per registry method, in registry order — the frame never
        # names a method; labels, order, and count all come from the engine
        for j, mr in enumerate(methods):
            ws.cell(row=r, column=2 + j, value=mr.label).font = Font(bold=True)
        r += 1
        hr = lambda k: f"N({self.hist_ref('balance', k)})"
        ws.cell(row=r, column=1, value="Enterprise value")
        for j, mr in enumerate(methods):
            c = ws.cell(row=r, column=2 + j, value=ev_formula[mr.id])
            c.number_format = FMT_M
        self.val["br_ev"] = f"Valuation!$B${r}"
        self.map["val:br_ev"] = ("Valuation", f"B{r}")
        ev_row = r
        r += 1
        items = [
            ("+ Excess cash (above operating floor)", "excess",
             (f"=MAX(0,{hr('cash_and_equivalents')}"
             f"+{hr('short_term_investments')}-cash_floor_pct"
             f"*N({self.hist_ref('income', 'revenue')}))")),
            ("+ Long-term investments", "lti", f"={hr('long_term_investments')}"),
            ("− Gross debt", "gd", f"=-{hr('gross_debt')}"),
            ("− Noncontrolling interest", "nci",
             f"=-{hr('noncontrolling_interest')}"),
            ("− Preferred equity", "pref", f"=-{hr('preferred_equity')}"),
            ("− Temporary equity", "temp", f"=-{hr('temporary_equity')}"),
            ("− Pension × (1 − marginal)", "pension",
             f"=-{hr('pension_liability')}*(1-marginal_tax)"),
        ]
        bridge_rows = []
        for label, key, formula in items:
            ws.cell(row=r, column=1, value=label)
            for j, _mr in enumerate(methods):
                c = ws.cell(row=r, column=2 + j,
                            value=formula if j == 0 else f"=$B${r}")
                c.number_format = FMT_M
            self.val[f"br_{key}"] = f"Valuation!$B${r}"
            self.map[f"val:br_{key}"] = ("Valuation", f"B{r}")
            bridge_rows.append(r)
            r += 1
        adj = "+".join(f"$B${row}" for row in bridge_rows)
        r = self._vrow(ws, r, "Bridge adjustment (Σ items — used by the "
                              "sensitivity grids)", "bridge_adj", f"={adj}")
        ws.cell(row=r, column=1, value="Equity value").font = Font(bold=True)
        for j, mr in enumerate(methods):
            col = get_column_letter(2 + j)
            ev_ref = f"${col}${ev_row}"
            c = ws.cell(
                row=r, column=2 + j,
                value=f"=IF(ISNUMBER({ev_ref}),{ev_ref}+{V['bridge_adj']},"
                      f"{ev_ref})")
            c.number_format = FMT_M
            c.font = Font(bold=True)
            self.val[f"equity_{mr.id}"] = f"Valuation!${col}${r}"
            self.map[f"val:equity_{mr.id}"] = ("Valuation", f"{col}{r}")
        r += 1
        ws.cell(row=r, column=1, value="VALUE PER SHARE").font = Font(bold=True)
        for j, mr in enumerate(methods):
            col = get_column_letter(2 + j)
            eq = V[f"equity_{mr.id}"]
            c = ws.cell(row=r, column=2 + j,
                        value=f"=IF(ISNUMBER({eq}),{eq}/share_count,{eq})")
            c.number_format = FMT_PS
            c.font = Font(bold=True)
            self.val[f"ps_{mr.id}"] = f"Valuation!${col}${r}"
            self.map[f"val:ps_{mr.id}"] = ("Valuation", f"{col}{r}")
        r += 2

        # ── value of growth: DCF (Gordon) − EPV — the comparison EPV exists
        #    for; the inverted case is labeled, never a negative number ──────
        self._header(ws, r, "VALUE OF GROWTH (DCF − EPV)", span=8)
        r += 1
        ps_g, ps_e = V["ps_gordon"], V["ps_epv"]
        r = self._vrow(ws, r, "Per share (Gordon DCF − EPV)", "growth_ps",
                       f"=IF(AND(ISNUMBER({ps_g}),ISNUMBER({ps_e})),"
                       f'{ps_g}-{ps_e},"n/a")', FMT_PS)
        r = self._vrow(ws, r, "As share of the DCF value", "growth_share",
                       f"=IF(AND(ISNUMBER({V['growth_ps']}),{ps_g}>0),"
                       f'{V["growth_ps"]}/{ps_g},"n/a")', FMT_PCT)
        r = self._vrow(ws, r, "Reading", "growth_state",
                       f"=IF(ISNUMBER({V['growth_ps']}),"
                       f'IF({V["growth_ps"]}<0,"growth is value-destructive '
                       'at these assumptions (EPV > DCF: returns below the '
                       'cost of capital, or shrinkage)","share of the DCF '
                       'value resting on growth beyond today\'s earnings '
                       'power"),"n/a")', "General")

    # ── Sensitivity ─────────────────────────────────────────────────────────
    def write_sensitivity(self) -> None:
        ws = self.ws_s
        ws.column_dimensions["A"].width = 34
        for i in range(1, 10):
            ws.column_dimensions[get_column_letter(1 + i)].width = 13
        V = self.val
        ws["A1"] = ("Sensitivity — live grids re-centering on the current "
                    "assumptions. The WACC × g grid re-projects per g column "
                    "(the growth path fades INTO terminal g — engine "
                    "semantics); helper blocks below are that re-projection, "
                    "kept visible on purpose.")
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(wrap_text=True)
        offsets = GRID_OFFSETS                # WACC rows / multiple cols
        r = 3

        if "wacc_x_g" in self.m.sensitivity:
            self._header(ws, r, "VALUE PER SHARE — WACC × terminal g", span=10)
            r += 1
            head = r
            ws.cell(row=head, column=1, value="WACC \\ g")
            for j, off in enumerate(G_OFFSETS):
                c = ws.cell(row=head, column=2 + j,
                            value=f"=terminal_growth+({off})*{G_STEP}")
                c.number_format = FMT_PCT
                c.font = Font(bold=True)
            helper_start = head + 8
            n6_row, ufcf_rows = self._g_helpers(ws, helper_start, head)
            last = self._mcol(self.horizon)
            for i, off in enumerate(offsets):
                wr = head + 1 + i
                wc = ws.cell(row=wr, column=1,
                             value=f"={V['wacc']}+({off})*{WACC_STEP}")
                wc.number_format = FMT_PCT
                wc.font = Font(bold=True)
                for j in range(len(G_OFFSETS)):
                    g = f"{get_column_letter(2 + j)}${head}"
                    w = f"$A{wr}"
                    u = ufcf_rows[j]
                    n6 = f"{get_column_letter(2 + j)}${n6_row}"
                    roic = (f'IF(OR(terminal_roic="",terminal_roic<={g}),'
                            f"{w},IF(terminal_roic_fade,"
                            f"(terminal_roic+{w})/2,terminal_roic))")
                    formula = (
                        f'=IF({g}>={w},"—",'
                        f"(SUMPRODUCT($C${u}:${last}${u},POWER(1+{w},"
                        f"-{self.t_rng}))"
                        f"+{n6}*(1-{g}/{roic})/({w}-{g})"
                        f"*POWER(1+{w},-{V['exp_gordon']})"
                        f"+{V['bridge_adj']})/share_count)")
                    c = ws.cell(row=wr, column=2 + j, value=formula)
                    c.number_format = FMT_PS
                    self.map[f"sens:g:{i}:{j}"] = (
                        "Sensitivity", f"{get_column_letter(2 + j)}{wr}")
            r = helper_start + len(G_OFFSETS) * 14 + 2
        else:
            ws.cell(row=r, column=1,
                    value="WACC × g grid unavailable — the Gordon leg has a "
                          "negative terminal anchor (see Valuation).")
            ws.cell(row=r, column=1).font = Font(color=RED)
            r += 2

        if "wacc_x_multiple" in self.m.sensitivity:
            self._header(ws, r, "VALUE PER SHARE — WACC × exit multiple "
                                "(base projection)", span=7)
            r += 1
            head = r
            ws.cell(row=head, column=1, value="WACC \\ multiple")
            for j, off in enumerate(offsets):
                c = ws.cell(row=head, column=2 + j,
                            value=f"=exit_multiple+({off})*1")
                c.number_format = FMT_X
                c.font = Font(bold=True)
            for i, off in enumerate(offsets):
                wr = head + 1 + i
                wc = ws.cell(row=wr, column=1,
                             value=f"={V['wacc']}+({off})*{WACC_STEP}")
                wc.number_format = FMT_PCT
                wc.font = Font(bold=True)
                for j in range(5):
                    mlt = f"{get_column_letter(2 + j)}${head}"
                    w = f"$A{wr}"
                    formula = (
                        f"=(SUMPRODUCT({self.ufcf_rng},POWER(1+{w},"
                        f"-{self.t_rng}))"
                        f"+{mlt}*{V['ebitda5']}*POWER(1+{w},-{V['tn']})"
                        f"+{V['bridge_adj']})/share_count")
                    c = ws.cell(row=wr, column=2 + j, value=formula)
                    c.number_format = FMT_PS
                    self.map[f"sens:mult:{i}:{j}"] = (
                        "Sensitivity", f"{get_column_letter(2 + j)}{wr}")
        elif "wacc_x_g" in self.m.sensitivity:
            ws.cell(row=r, column=1,
                    value="WACC × multiple grid unavailable — no exit "
                          "multiple (FY0 EBITDA ≤ 0) or negative FY5 EBITDA.")
            ws.cell(row=r, column=1).font = Font(color=RED)

    def _g_helpers(self, ws, start: int, head: int) -> tuple[int, list[int]]:
        """One labeled re-projection block per g column (len(G_OFFSETS) of
        them). Returns the row of the NOPAT6 cells and the UFCF row per
        block."""
        R = self.model_row
        ratio_terms = ["rnd_pct", "sga_pct", "other_opex_pct",
                       "unclassified_costs_pct"]
        if self.by_function:
            ratio_terms.insert(0, "cogs_pct")
        ratios = "+".join(ratio_terms)
        ws.cell(row=start - 1, column=1,
                value="Helper — re-projection per terminal-g column "
                      "(UFCF depends on g through the growth fade)"
                ).font = Font(color=GRAY, italic=True)
        ufcf_rows: list[int] = []
        n6_row = 0
        for j in range(len(G_OFFSETS)):
            base = start + j * 14
            g = f"{get_column_letter(2 + j)}${head}"
            ws.cell(row=base, column=1,
                    value=f"g column {j + 1}").font = Font(bold=True)
            rows = {}
            for k, label in (("growth", "growth"), ("rev", "revenue"),
                             ("ebit", "EBIT"), ("capex", "capex"),
                             ("dep", "depreciation"),
                             ("amort", "intangible amortization"),
                             ("da", "D&A"), ("ppe", "PP&E"),
                             ("intang", "intangibles"), ("nwc", "NWC"),
                             ("dnwc", "ΔNWC"), ("ufcf", "UFCF")):
                rows[k] = base + 1 + len(rows)
                ws.cell(row=rows[k], column=1,
                        value=f"  {label}").font = Font(color=GRAY, size=9)
            for i in range(1, self.horizon + 1):
                col = self._mcol(i)
                prev = self._mcol(i - 1)
                # same live fade-shape branch as the Model sheet, at this
                # column's g (the grid re-projects per column, cosine included)
                ws.cell(row=rows["growth"], column=2 + i,
                        value=(f"=IF(fade_curved,{g}+"
                               f"(revenue_growth_fy1-{g})*"
                               f"(1+COS(PI()*{i - 1}/{self.horizon - 1}))/2,"
                               f"revenue_growth_fy1+{i - 1}/"
                               f"{self.horizon - 1}*({g}-revenue_growth_fy1))")
                        ).number_format = FMT_PCT
                rev_prev = (f"Model!$B${R['revenue']}" if i == 1
                            else f"{prev}{rows['rev']}")
                ws.cell(row=rows["rev"], column=2 + i,
                        value=f"={rev_prev}*(1+{col}{rows['growth']})"
                        ).number_format = FMT_M
                ws.cell(row=rows["ebit"], column=2 + i,
                        value=f"={col}{rows['rev']}*(1-({ratios}))"
                        ).number_format = FMT_M
                ws.cell(row=rows["capex"], column=2 + i,
                        value=(f"=IF(capex_fade,capex_pct+{i - 1}/"
                               f"{self.horizon - 1}*"
                               f"(capex_terminal_pct-capex_pct),capex_pct)"
                               f"*{col}{rows['rev']}")
                        ).number_format = FMT_M
                ppe_prev = (f"Model!$B${R['ppe']}" if i == 1
                            else f"{prev}{rows['ppe']}")
                intang_prev = (f"Model!$B${R['intangibles']}" if i == 1
                               else f"{prev}{rows['intang']}")
                # split D&A basis, mirroring the Model sheet: depreciation
                # drives the PP&E roll (MIN = identity floor), amortization
                # runs the intangibles balance off
                dep_f = (f"=MIN(dep_pct_beginning_ppe*{ppe_prev},"
                         f"{ppe_prev}+{col}{rows['capex']})"
                         if self.da_on_ppe
                         else f"=da_pct_revenue*{col}{rows['rev']}")
                ws.cell(row=rows["dep"], column=2 + i,
                        value=dep_f).number_format = FMT_M
                amort_f = (f"=MIN(amort_pct_beginning_intangibles"
                           f"*{intang_prev},{intang_prev})"
                           if self.da_on_ppe
                           and self.a.has("amort_pct_beginning_intangibles")
                           else 0)
                ws.cell(row=rows["amort"], column=2 + i,
                        value=amort_f).number_format = FMT_M
                ws.cell(row=rows["da"], column=2 + i,
                        value=f"={col}{rows['dep']}+{col}{rows['amort']}"
                        ).number_format = FMT_M
                ws.cell(row=rows["ppe"], column=2 + i,
                        value=f"={ppe_prev}+{col}{rows['capex']}"
                              f"-{col}{rows['dep']}").number_format = FMT_M
                ws.cell(row=rows["intang"], column=2 + i,
                        value=f"={intang_prev}-{col}{rows['amort']}"
                        ).number_format = FMT_M
                basis = (f"cogs_pct*{col}{rows['rev']}" if self.by_function
                         else f"({col}{rows['rev']}-{col}{rows['ebit']})")
                ws.cell(row=rows["nwc"], column=2 + i,
                        value=f"={col}{rows['rev']}*(dso/365+oca_pct"
                              f"-accrued_pct-ocl_pct-defrev_pct)"
                              f"+{basis}*(dio-dpo)/365").number_format = FMT_M
                nwc_prev = (f"Model!$B${R['nwc']}" if i == 1
                            else f"{prev}{rows['nwc']}")
                ws.cell(row=rows["dnwc"], column=2 + i,
                        value=f"={col}{rows['nwc']}-{nwc_prev}"
                        ).number_format = FMT_M
                ws.cell(row=rows["ufcf"], column=2 + i,
                        value=f"={col}{rows['ebit']}"
                              f"*(1-{self.model_ref('tax_rate', i)})"
                              f"+{col}{rows['da']}"
                              f"+IF(sbc_addback,sbc_pct*{col}{rows['rev']},0)"
                              f"-{col}{rows['capex']}-{col}{rows['dnwc']}"
                        ).number_format = FMT_M
            ufcf_rows.append(rows["ufcf"])
        n6_row = start + len(G_OFFSETS) * 14
        ws.cell(row=n6_row, column=1,
                value="NOPAT(N+1) per g column").font = Font(color=GRAY, size=9)
        for j in range(len(G_OFFSETS)):
            g = f"{get_column_letter(2 + j)}${head}"
            ufcf_row = ufcf_rows[j]
            ebit5 = f"{self._mcol(self.horizon)}{ufcf_row - 9}"  # ebit row
            ws.cell(row=n6_row, column=2 + j,
                    value=f"={ebit5}*(1+{g})*(1-marginal_tax)"
                    ).number_format = FMT_M
        return n6_row, ufcf_rows

    # ── Methodology ─────────────────────────────────────────────────────────
    def write_methodology(self) -> None:
        ws = self.ws_meth
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 130
        ws["A1"] = "Methodology — rendered from engine/methodology.yaml"
        ws["A1"].font = Font(bold=True, size=12)
        engine_dir = Path(__file__).parent.parent / "engine"
        doc = yaml.safe_load((engine_dir / "methodology.yaml").read_text())
        r = 3
        for entry in doc.get("conventions", []):
            ws.cell(row=r, column=1, value=entry.get("label", entry["id"])
                    ).font = Font(bold=True)
            ws.cell(row=r, column=2, value=f"Default: {entry.get('default', '')}")
            r += 1
            ws.cell(row=r, column=2,
                    value=f"Derivation: {entry.get('derivation', '')}")
            r += 1
            trade = " ".join(str(entry.get("tradeoff", "")).split())
            if trade:
                c = ws.cell(row=r, column=2, value=f"Tradeoff: {trade}")
                c.font = Font(color=GRAY)
                r += 1
            r += 1
        self._header(ws, r, "ASSUMPTION PRESETS (engine/presets.yaml)", span=2)
        r += 1
        pdoc = yaml.safe_load((engine_dir / "presets.yaml").read_text())
        for p in pdoc.get("presets", []):
            ws.cell(row=r, column=1, value=p.get("title", p["name"])
                    ).font = Font(bold=True)
            ws.cell(row=r, column=2,
                    value=" ".join(str(p.get("rationale", "")).split()))
            r += 1
            for fname, spec in (p.get("fields") or {}).items():
                rule = spec.get("rule") or spec.get("value") \
                    or f"{spec.get('solver')} -> {spec.get('target')}"
                c = ws.cell(row=r, column=2,
                            value=f"{fname}: {spec['form']} — "
                                  + " ".join(str(rule).split()))
                c.font = Font(color=GRAY, size=9)
                r += 1
            r += 1

    # ── Cover ───────────────────────────────────────────────────────────────
    def write_cover(self) -> None:
        ws = self.ws_cover
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 100
        m, V = self.m, self.val
        c = ws["A1"]
        c.value = f"{self.h.company.name} ({m.ticker}) — DCF valuation"
        c.font = Font(bold=True, size=14)
        ws["A2"] = (f"{self.h.company.sic_description} · FYE anchor "
                    f"{self.h.company.fye_anchor} · history "
                    f"FY{self.h.periods[0].fiscal_year}–"
                    f"FY{self.h.periods[-1].fiscal_year} "
                    f"({self.a.cost_structure})")
        ws["A3"] = ("Convention: blue = hardcoded input/actual, black = live "
                    "formula, green = cross-sheet link. Change any Assumptions "
                    "cell and everything recalculates.")
        ws["A3"].font = Font(color=GRAY, size=9)

        r = 5
        rows = [
            ("valuation_date", "Valuation date", "=valuation_date", FMT_DATE),
            ("price", "Market price", "=market_price", FMT_PS),
        ]
        # headline per-share values iterate the registry (labels and order
        # from the engine — Cover never names a method)
        for mr in sorted(self.m.methods, key=lambda mr: mr.order):
            ps = V[f"ps_{mr.id}"]
            rows += [
                (f"ps_{mr.id}", f"{mr.label} (per share)", f"={ps}",
                 FMT_PS),
                (f"vs_price_{mr.id}", "  vs price",
                 f'=IF(ISNUMBER({ps}),{ps}/market_price-1,"—")', FMT_PCT),
            ]
        rows += [
            ("growth_ps", "Value of growth (DCF − EPV, per share)",
             f"={V['growth_ps']}", FMT_PS),
            ("wacc", "WACC", f"={V['wacc']}", FMT_PCT),
            ("implied_mult", "Implied exit multiple (from your g)",
             f"={V['implied_mult']}", FMT_X),
            ("implied_g", "Implied terminal g (from your multiple)",
             f"={V['implied_g']}", FMT_PCT),
        ]
        for key, label, formula, fmt in rows:
            ws.cell(row=r, column=1, value=label)
            cell = ws.cell(row=r, column=2, value=formula)
            cell.number_format = fmt
            cell.font = Font(bold=True)
            self.map[f"cover:{key}"] = ("Cover", f"B{r}")
            r += 1
        r += 1

        if self.preset is not None or self.a.active_preset:
            name = self.preset.title if self.preset else self.a.active_preset
            rationale = " ".join(self.preset.rationale.split()) \
                if self.preset else "see Methodology sheet"
            ws.cell(row=r, column=1, value="Active preset").font = Font(bold=True)
            ws.cell(row=r, column=2, value=name).font = Font(bold=True)
            ws.cell(row=r, column=3, value=rationale)
        else:
            ws.cell(row=r, column=1, value="Active preset").font = Font(bold=True)
            ws.cell(row=r, column=2, value="derived (defaults)")
        r += 2

        self._header(ws, r, "LIVE CHECKS", span=3)
        r += 1
        checks = [
            ("Balance sheet ties every projected year",
             f"=Model!C{self.model_row['ok_bs']}"),
            ("Cash flow ties to Δcash every projected year",
             f"=Model!C{self.model_row['ok_cf']}"),
            ("Terminal g below WACC",
             f'=IF(terminal_growth<{V["wacc"]},"OK","FAIL — g ≥ WACC")'),
            ("Reinvestment rate below 1",
             f'=IF({V["rr"]}<1,"OK","FAIL — RR ≥ 1")'),
        ]
        for label, formula in checks:
            ws.cell(row=r, column=1, value=label)
            cell = ws.cell(row=r, column=2, value=formula)
            cell.font = Font(bold=True)
            self.map[f"cover:check:{label.split()[0].lower()}"] = ("Cover", f"B{r}")
            r += 1
        r += 1

        warnings = self._warning_lines()
        self._header(ws, r, f"WARNINGS ({len(warnings)})", span=3)
        r += 1
        for line in warnings or ["none"]:
            cell = ws.cell(row=r, column=1, value=line)
            cell.font = Font(color=RED, size=9)
            r += 1
        r += 1
        ws.cell(row=r, column=1,
                value=f"Historical validation: {self.h.validation.overall} · "
                      "projection checks at generation: "
                      + "  ".join(f"{c.check_id}={c.status}"
                                  for c in m.checks.results)
                ).font = Font(color=GRAY, size=9)
        r += 1
        ws.cell(row=r, column=1,
                value=f"Price {self.m.market.price.staleness} · risk-free "
                      f"{self.m.market.risk_free.staleness} · generated by "
                      "Ticker-to-Model; every convention on the Methodology "
                      "sheet").font = Font(color=GRAY, size=9)

    def _warning_lines(self) -> list[str]:
        lines = []
        unmapped: dict[str, list[int]] = {}
        for w in self.h.warnings:
            if w.code == "unmapped_item" and w.item:
                unmapped.setdefault(w.item, []).append(w.fiscal_year)
        for origin, code, message in self.m.all_warnings():
            if origin == "ingest" and code == "unmapped_item":
                continue
            lines.append(f"[{origin}:{code}] {message}")
        for item, ys in unmapped.items():
            years = [y for y in ys if y] or [0]
            span = (f"FY{min(years)}–FY{max(years)}" if len(years) > 1
                    else f"FY{years[0]}")
            lines.append(f"[ingest:unmapped_item] {item} unmapped in {span}; "
                         "treated as 0 per the documented per-item rule")
        return lines


def write_workbook(model: ModelResult, path: str | Path,
                   preset: Preset | None = None) -> dict[str, tuple[str, str]]:
    """Write the formula-driven workbook; returns the cell map used by the
    round-trip parity test ({logical key: (sheet, coordinate)})."""
    w = _Writer(model, preset)
    w.build()
    w.wb.save(path)
    return w.map
